"""
Export Excel pour les résultats du Monte Carlo d'estimation de covariance.

Ce fichier gère la persistance et la reprise des simulations Monte Carlo.
Il exporte les résultats dans un fichier Excel multi-onglets avec un mécanisme
de checkpoint (skip automatique des scénarios déjà calculés au redémarrage).

Deux branches d'export coexistent dans ce fichier :
- Branche économique (MonteCarloExporter) : backtest TE-min, métriques par scénario.
- Branche statistique (StatMonteCarloExporter) : pertes Frobenius / spectral / Stein / précision.

Structure du fichier Excel économique (3 onglets)
---------------------------------------------------
- ts_results : une ligne par (scénario x année) avec rendements et TE ex-post.
- summary : une ligne par scénario avec toutes les métriques agrégées.
- run_log : une ligne par scénario avec statut, durée et erreur éventuelle.

Structure du fichier Excel statistique (2 onglets)
----------------------------------------------------
- stat_summary : une ligne par (modèle x seed) avec les pertes matricielles.
- stat_log : statut, durée, erreur.

Classes
-------
ScenarioKey :
    Identifiant unique d'un scénario Monte Carlo économique (modèle, rolling, seed, etc.).
ScenarioResult :
    Résultat complet d'un scénario économique (séries, métriques, statut).
StatScenarioKey :
    Identifiant unique d'un scénario de simulation statistique.
StatScenarioResult :
    Résultat d'un scénario statistique (pertes matricielles, statut).
MonteCarloExporter :
    Interface publique pour l'export économique avec checkpoint et flush bufferisé.
StatMonteCarloExporter :
    Interface publique pour l'export statistique avec checkpoint et flush bufferisé.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, Optional, Set, Any, List

import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)



# Constantes : noms des feuilles Excel
SHEET_TS  = "ts_results"   # séries temporelles par année
SHEET_SUMMARY = "summary"      # métriques agrégées par scénario
SHEET_LOG  = "run_log"      # statut et durée par scénario



# Dataclass : clé d'un scénario économique
@dataclass(frozen=True)
class ScenarioKey:
    """
    Classe contenant les 6 paramètres qui identifient de façon unique un scénario
    Monte Carlo économique. Utilisée comme clé de checkpoint et comme identifiant
    dans le fichier Excel.

    Attributes
    ----------
    model_name : str
        Nom du modèle de covariance (ex: 'Rolling', 'EWMA', 'LW').
    rolling : int
        Taille de la fenêtre glissante utilisée pour l'estimation (en jours).
    exclude_frac : float
        Fraction d'actifs exclus de l'univers investissable (ex: 0.20 = 20% exclus).
    seed : int
        Graine aléatoire du scénario (pour la reproductibilité).
    rebal_freq : str
        Fréquence de rebalancement ('M', 'Q', 'W', etc.).
    data_freq : str
        Fréquence des données ('daily' ou 'weekly').

    Methods
    -------
    to_dict() -> dict :
        Sérialise la clé en dictionnaire pour l'écriture dans Excel.
    """
    model_name:   str
    rolling:      int
    exclude_frac: float
    seed:         int
    rebal_freq:   str    
    data_freq:    str = "daily" 

    def to_dict(self) -> Dict[str, Any]:
        """Sérialise la clé en dictionnaire (pour l'écriture dans Excel)."""
        return asdict(self)

    def __str__(self) -> str:
        return (f"{self.model_name} | roll={self.rolling} | "f"excl={self.exclude_frac:.0%} | seed={self.seed} | " f"freq={self.rebal_freq}")



# Dataclass : résultat complet d'un scénario économique
@dataclass
class ScenarioResult:
    """
    Classe contenant l'ensemble des données produites par un scénario Monte Carlo
    économique : séries temporelles du backtest, métriques agrégées sur la période,
    pertes matricielles statistiques et statut d'exécution.

    Attributes
    ----------
    key : ScenarioKey
        Identifiant du scénario.
    port_returns : pd.Series or None
        Rendements journaliers du portefeuille.
    bench_returns : pd.Series or None
        Rendements journaliers du benchmark.
    active_returns : pd.Series or None
        Rendements actifs (port - bench), base de toutes les métriques TE.
    te_expost : pd.Series or None
        TE ex-post rolling 252j, annualisée.
    te_ann : float or None
        TE annualisée sur toute la période.
    te_daily : float or None
        TE journalière (écart-type des active returns).
    information_ratio : float or None
        Ratio rendement actif / TE (IR = mu_ann / te_ann).
    active_mean_ann : float or None
        Rendement actif annualisé moyen.
    port_cum / bench_cum : float or None
        Performance cumulée sur la période (portefeuille et benchmark).
    rel_cum_vs_bench : float or None
        Performance relative : port_cum - bench_cum.
    max_dd_active : float or None
        Maximum drawdown des active returns sur la période.
    skewness_active / kurtosis_active : float or None
        Moments d'ordre 3 et 4 de la distribution des active returns.
    te_stability : float or None
        Écart-type des TE annuelles (mesure de régularité de la TE).
    te_ann_bear / te_ann_bull : float or None
        TE annualisée sur les jours où le benchmark est négatif / positif.
    worst_year_active / best_year_active : float or None
        Pire / meilleure année de rendement actif.
    turnover_ann : float or None
        Turnover annualisé moyen (en fraction de NAV).
    te_per_unit_turnover : float or None
        Efficacité : TE / turnover (moins = plus efficace par unité de coût).
    frobenius / spectral / stein / precision : float or None
        Pertes matricielles de l'estimateur vs covariance vraie (branche stats).
    status : str
        'success' ou 'error'.
    error : str or None
        Message d'erreur en cas d'échec (tronqué à 250 caractères).
    duration : float or None
        Durée d'exécution du scénario en secondes.
    ann_factor : int
        Facteur d'annualisation (252 pour données journalières).
    """

    key: ScenarioKey

    # Séries temporelles (DatetimeIndex)
    port_returns:   Optional[pd.Series] = None
    bench_returns:  Optional[pd.Series] = None
    active_returns: Optional[pd.Series] = None
    te_expost:      Optional[pd.Series] = None   

    # Métriques éco agrégées
    te_ann:            Optional[float] = None
    te_daily:          Optional[float] = None
    information_ratio: Optional[float] = None
    active_mean_ann:   Optional[float] = None
    port_cum:          Optional[float] = None
    bench_cum:         Optional[float] = None
    rel_cum_vs_bench:  Optional[float] = None
    max_dd_active:         Optional[float] = None
    skewness_active:       Optional[float] = None
    kurtosis_active:       Optional[float] = None
    te_stability:          Optional[float] = None   
    te_ann_bear:           Optional[float] = None   
    te_ann_bull:           Optional[float] = None   
    worst_year_active:     Optional[float] = None
    best_year_active:      Optional[float] = None
    turnover_ann:          Optional[float] = None
    te_per_unit_turnover:  Optional[float] = None

    # Métriques stats agrégées
    frobenius: Optional[float] = None
    spectral:  Optional[float] = None
    stein:     Optional[float] = None
    precision: Optional[float] = None

    # Statut d'exécution
    status:   str            = "success"
    error:    Optional[str]  = None
    duration: Optional[float] = None   # secondes

    # facteur d'annualisation (252 jours de bourse par an)
    ann_factor: int = 252  



# Schéma des colonnes Excel (branche économique)

# Colonnes communes à tous les onglets, identifiant un scénario de façon unique
_KEY_COLS = ["model_name", "rolling", "exclude_frac", "seed", "rebal_freq", "data_freq"]

# Colonnes de la feuille ts_results : une ligne par année de backtest
_TS_COLS = _KEY_COLS + ["date", "port_return", "bench_return", "active_return", "te_expost","max_dd_active_yr", "skewness_active_yr", "kurtosis_active_yr",]

# Colonnes de la feuille summary : une ligne par scénario, toutes les métriques agrégées
_SUMMARY_COLS = _KEY_COLS + ["te_ann", "te_daily", "information_ratio", "active_mean_ann","port_cum", "bench_cum", "rel_cum_vs_bench",
                             "max_dd_active", "skewness_active", "kurtosis_active","te_stability", "te_ann_bear", "te_ann_bull",
                             "worst_year_active", "best_year_active","turnover_ann", "te_per_unit_turnover","frobenius", "spectral", "stein", "precision",]

# Colonnes de la feuille run_log : statut, durée et erreur éventuelle
_LOG_COLS = _KEY_COLS + ["status", "duration_s", "error"]



# Classe interne : mise en forme des feuilles Excel
class _Styler:
    """
    Classe contenant les méthodes de mise en forme appliquées aux feuilles Excel.
    Gère le style du header (fond bleu marine, texte blanc gras) et les largeurs
    de colonnes prédéfinies pour les colonnes connues.

    Methods
    -------
    apply(ws, cols) -> None :
        Applique le style header et les largeurs de colonnes sur une feuille.
    """

    # Fond bleu marine pour le header
    FILL  = PatternFill("solid", fgColor="1E3A5F")  

    # Texte blanc gras pour le header
    FONT  = Font(color="FFFFFF", bold=True, size=9)  

    ALIGN = Alignment(horizontal="center", vertical="center")

    # Largeurs de colonnes prédéfinies pour les colonnes connues
    WIDTHS: Dict[str, int] = {"model_name": 18, "rolling": 10, "exclude_frac": 14, "seed": 8, "rebal_freq": 10, "date": 14,"status": 10, "error": 35, "duration_s": 12,}

    # Largeur par défaut pour les colonnes non listées dans WIDTHS
    DEFAULT = 14  

    @classmethod
    def apply(cls, ws, cols: List[str]) -> None:
        """
        Applique le style header et les largeurs de colonnes sur une feuille.

        Parameters
        ----------
        ws : openpyxl worksheet
            Feuille à styler.
        cols : list[str]
            Noms des colonnes dans l'ordre (pour calculer les largeurs).
        """
        # Applique le style bleu sur chaque cellule de la première ligne (header)
        for cell in ws[1]:  
            cell.fill  = cls.FILL
            cell.font  = cls.FONT
            cell.alignment = cls.ALIGN

        # Définit la largeur de chaque colonne selon le dictionnaire WIDTHS
        for i, col in enumerate(cols, 1):  
            ws.column_dimensions[get_column_letter(i)].width = (cls.WIDTHS.get(col, cls.DEFAULT))

        # Gèle la première ligne pour qu'elle reste visible au défilement
        ws.freeze_panes = "A2"  



# Classe interne : initialisation du fichier Excel
class _Initializer:
    """
    Classe contenant les méthodes de création et de vérification du fichier Excel.
    Si le fichier n'existe pas, le crée avec les 3 feuilles et leurs headers.
    Si le fichier existe, vérifie que les 3 feuilles sont présentes et les ajoute
    si manquantes.

    Methods
    -------
    init(path) -> None :
        Point d'entrée : création ou vérification selon l'existence du fichier.
    """

    # Mapping nom de feuille vers ses colonnes
    SHEETS = {SHEET_TS: _TS_COLS, SHEET_SUMMARY: _SUMMARY_COLS, SHEET_LOG: _LOG_COLS,}

    @classmethod
    def init(cls, path: Path) -> None:
        if path.exists():

            # Le fichier existe : on vérifie que les feuilles attendues sont présentes
            cls._ensure_sheets(path) 

        else:

            # Le fichier n'existe pas : on le crée de zéro avec toutes les feuilles
            cls._create(path) 

    @classmethod
    def _create(cls, path: Path) -> None:
        """Crée un nouveau fichier Excel avec les 3 feuilles formatées."""
        wb = Workbook()

        # Supprime la feuille vide créée automatiquement par openpyxl
        wb.remove(wb.active)  

        for name, cols in cls.SHEETS.items():
            ws = wb.create_sheet(name)

            # Écrit la ligne de header avec les noms de colonnes
            ws.append(cols)  

            # Applique le style bleu sur le header de cette feuille
            _Styler.apply(ws, cols)  

        # Crée les répertoires parents si le chemin de sortie n'existe pas encore
        path.parent.mkdir(parents=True, exist_ok=True)  

        wb.save(path)
        logger.info(f"[Export] Fichier créé : {path}")

    @classmethod
    def _ensure_sheets(cls, path: Path) -> None:
        """Ajoute les feuilles manquantes dans un fichier existant."""
        wb = load_workbook(path)
        changed = False
        for name, cols in cls.SHEETS.items():
            if name not in wb.sheetnames:
                ws = wb.create_sheet(name)
                ws.append(cols)
                _Styler.apply(ws, cols)
                changed = True
                
        # Sauvegarde uniquement si au moins une feuille a été ajoutée
        if changed:
            wb.save(path)  



# Classe interne : checkpoint des scénarios déjà calculés
class _Checkpoint:
    """
    Classe contenant les méthodes de lecture des scénarios déjà présents dans
    le fichier Excel. Permet de reprendre un Monte Carlo interrompu sans
    recalculer les scénarios déjà terminés.

    Methods
    -------
    load(path) -> set[ScenarioKey] :
        Charge les clés existantes depuis la feuille summary.
    """

    @staticmethod
    def load(path: Path) -> Set[ScenarioKey]:
        """
        Charge les clés existantes depuis la feuille summary.

        Returns
        -------
        set[ScenarioKey]
            Ensemble des clés déjà calculées. Vide si fichier absent ou illisible.
        """
        if not path.exists():
            return set()
        try:
            df = pd.read_excel(path, sheet_name=SHEET_SUMMARY, engine="openpyxl")
            return _Checkpoint._parse(df)
        except Exception as e:
            logger.warning(f"[Checkpoint] Lecture impossible : {e}")
            return set()

    @staticmethod
    def _parse(df: pd.DataFrame) -> Set[ScenarioKey]:
        """Convertit les lignes du DataFrame en ScenarioKey. Ignore les lignes malformées."""
        keys: Set[ScenarioKey] = set()

        # Si les colonnes attendues ne sont pas toutes présentes, le fichier est incompatible
        if not set(_KEY_COLS).issubset(df.columns):
            return keys

        for _, r in df.iterrows():
            try:
                keys.add(ScenarioKey(
                    model_name   = str(r["model_name"]),
                    rolling      = int(r["rolling"]),
                    exclude_frac = float(r["exclude_frac"]),
                    seed         = int(r["seed"]),
                    rebal_freq   = str(r["rebal_freq"]),
                    # data_freq absent dans les anciens fichiers : on met "daily" par défaut
                    data_freq    = str(r.get("data_freq", "daily"))
                ))
            except Exception:
                # La ligne est malformée, on passe à la suivante sans planter
                continue
        return keys


# Classe interne : écriture dans Excel
class _Writer:
    """
    Classe contenant les méthodes d'écriture des résultats dans le fichier Excel.
    Deux modes coexistent : écriture atomique (ouvre, écrit, ferme) et écriture
    sur un workbook déjà ouvert en mémoire (pour les flush bufferisés).

    Methods
    -------
    write_ts(path, result, ann_factor) -> None :
        Écrit ts_results (une ligne par année) en écriture atomique.
    write_ts_wb(wb, result, ann_factor) -> None :
        Écrit ts_results sur un workbook ouvert (pour flush bufferisé).
    write_summary(path, result) -> None :
        Écrit 1 ligne dans summary en écriture atomique.
    write_log(path, result) -> None :
        Écrit 1 ligne dans run_log en écriture atomique.
    """

    @staticmethod
    def _append(path: Path, sheet: str, cols: List[str], values: Dict[str, Any],ann_factor : int = 252) -> None:
        """Ouvre le fichier, écrit une ligne, sauvegarde et ferme (écriture atomique)."""
        wb = load_workbook(path)

        # Extrait les valeurs dans l'ordre des colonnes pour construire la ligne
        wb[sheet].append([values.get(c) for c in cols])
        wb.save(path)

    @staticmethod
    def _append_wb(wb, sheet: str, cols: List[str], values: Dict[str, Any]) -> None:
        """Écrit une ligne sur un workbook déjà ouvert (pas de save, pour les flush bufferisés)."""
        wb[sheet].append([values.get(c) for c in cols])

    @classmethod
    def write_ts_wb(cls, wb, result: ScenarioResult, ann_factor: int) -> None:
        """
        Écrit les données ts_results (une ligne par année) sur un workbook ouvert.

        Agrège les rendements journaliers par année, calcule les métriques annuelles
        (max drawdown, skewness, kurtosis actif) et les écrit dans la feuille ts_results.
        """
        # Si les séries ne sont pas disponibles, rien à écrire
        if result.active_returns is None:
            return
        key_vals = result.key.to_dict()
        dates    = result.active_returns.index

        # Réindexe les séries sur l'index des active returns pour aligner les dates
        port_s   = _reindex_or_nan(result.port_returns,  dates)
        bench_s  = _reindex_or_nan(result.bench_returns, dates)
        te_s     = _reindex_or_nan(result.te_expost,     dates)

        years        = dates.year

        # Calcule les performances annuelles cumulées par groupby sur l'année
        port_yearly  = (1 + port_s).groupby(years).prod() - 1
        bench_yearly = (1 + bench_s).groupby(years).prod() - 1
        active_yearly = port_yearly - bench_yearly

        # TE ex-post annualisée par année : écart-type intra-année * sqrt(ann_factor)
        te_yearly    = te_s.groupby(years).std() * np.sqrt(ann_factor)

        ws = wb[SHEET_TS]
        for year in port_yearly.index:
            mask_yr   = dates.year == year

            # Active returns de l'année courante, sans NaN
            active_yr = result.active_returns.loc[mask_yr].dropna()
            if len(active_yr) > 0:
                cum_yr    = (1 + active_yr).cumprod()

                # Max drawdown de l'année : pire chute depuis le sommet
                max_dd_yr = _sf((((cum_yr - cum_yr.cummax()) / cum_yr.cummax())).min())
            else:
                max_dd_yr = None
            row = {
                **key_vals,
                "date":               year,
                "port_return":        _sf(port_yearly[year]),
                "bench_return":       _sf(bench_yearly[year]),
                "active_return":      _sf(active_yearly[year]),
                "te_expost":          _sf(te_yearly.get(year)),
                "max_dd_active_yr":   max_dd_yr,

                # Skewness et kurtosis calculés uniquement si assez d'observations
                "skewness_active_yr": _sf(float(active_yr.skew()))       if len(active_yr) > 3 else None,
                "kurtosis_active_yr": _sf(float(active_yr.kurtosis()))   if len(active_yr) > 3 else None,
            }
            ws.append([row.get(c) for c in _TS_COLS])

    @classmethod
    def write_summary_wb(cls, wb, result: ScenarioResult) -> None:
        """Écrit la ligne summary (métriques agrégées) sur un workbook ouvert."""
        values = {
            **result.key.to_dict(),
            "te_ann": result.te_ann, "te_daily": result.te_daily,
            "information_ratio": result.information_ratio,
            "active_mean_ann": result.active_mean_ann,
            "port_cum": result.port_cum, "bench_cum": result.bench_cum,
            "rel_cum_vs_bench": result.rel_cum_vs_bench,
            "max_dd_active": result.max_dd_active,
            "skewness_active": result.skewness_active,
            "kurtosis_active": result.kurtosis_active,
            "te_stability": result.te_stability,
            "te_ann_bear": result.te_ann_bear, "te_ann_bull": result.te_ann_bull,
            "worst_year_active": result.worst_year_active,
            "best_year_active": result.best_year_active,
            "turnover_ann": result.turnover_ann,
            "te_per_unit_turnover": result.te_per_unit_turnover,
            "frobenius": result.frobenius, "spectral": result.spectral,
            "stein": result.stein, "precision": result.precision,
        }
        cls._append_wb(wb, SHEET_SUMMARY, _SUMMARY_COLS, values)

    @classmethod
    def write_log_wb(cls, wb, result: ScenarioResult) -> None:
        """Écrit la ligne run_log (statut, durée, erreur) sur un workbook ouvert."""
        values = { **result.key.to_dict(), "status": result.status, "duration_s": round(result.duration, 1) if result.duration else None, "error": result.error,}
        cls._append_wb(wb, SHEET_LOG, _LOG_COLS, values)

    @classmethod
    def write_ts(cls, path: Path, result: ScenarioResult,ann_factor) -> None:
        """N lignes dans ts_results (une par date)."""
        if result.active_returns is None:
            return
        key_vals = result.key.to_dict()
        dates    = result.active_returns.index

        port_s  = _reindex_or_nan(result.port_returns,  dates)
        bench_s = _reindex_or_nan(result.bench_returns, dates)
        te_s    = _reindex_or_nan(result.te_expost,     dates)

        years = dates.year

        # Performances annuelles cumulées pour chaque année présente dans les données
        port_yearly  = (1 + port_s).groupby(years).prod() - 1
        bench_yearly = (1 + bench_s).groupby(years).prod() - 1
        active_yearly = port_yearly - bench_yearly

        # TE ex-post annualisé par année : écart-type * sqrt(ann_factor)
        te_yearly = te_s.groupby(years).std() * np.sqrt(ann_factor)

        wb = load_workbook(path)
        ws = wb[SHEET_TS]
        for year in port_yearly.index:

            # Active returns de l'année pour calculer les métriques intra-année
            mask_yr   = dates.year == year
            active_yr = result.active_returns.loc[mask_yr].dropna()

            # Max drawdown actif de l'année
            if len(active_yr) > 0:
                cum_yr   = (1 + active_yr).cumprod()
                roll_max = cum_yr.cummax()
                dd_yr    = (cum_yr - roll_max) / roll_max
                max_dd_yr = _sf(dd_yr.min())
            else:
                max_dd_yr = None

            # Skewness et kurtosis calculés uniquement si assez d'observations
            skew_yr = _sf(float(active_yr.skew()))   if len(active_yr) > 3 else None
            kurt_yr = _sf(float(active_yr.kurtosis())) if len(active_yr) > 3 else None

            row = {
                **key_vals,
                "date":                year,
                "port_return":         _sf(port_yearly[year]),
                "bench_return":        _sf(bench_yearly[year]),
                "active_return":       _sf(active_yearly[year]),
                "te_expost":           _sf(te_yearly.get(year)),
                "max_dd_active_yr":    max_dd_yr,
                "skewness_active_yr":  skew_yr,
                "kurtosis_active_yr":  kurt_yr,
            }

            ws.append([row.get(c) for c in _TS_COLS])
        wb.save(path)

    @classmethod
    def write_summary(cls, path: Path, result: ScenarioResult) -> None:
        """1 ligne dans summary."""
        values = {
            **result.key.to_dict(),
            "te_ann":            result.te_ann,
            "te_daily":          result.te_daily,
            "information_ratio": result.information_ratio,
            "active_mean_ann":   result.active_mean_ann,
            "port_cum":          result.port_cum,
            "bench_cum":         result.bench_cum,
            "rel_cum_vs_bench":  result.rel_cum_vs_bench,
            "max_dd_active":        result.max_dd_active,
            "skewness_active":      result.skewness_active,
            "kurtosis_active":      result.kurtosis_active,
            "te_stability":         result.te_stability,
            "te_ann_bear":          result.te_ann_bear,
            "te_ann_bull":          result.te_ann_bull,
            "worst_year_active":    result.worst_year_active,
            "best_year_active":     result.best_year_active,
            "turnover_ann":         result.turnover_ann,
            "te_per_unit_turnover": result.te_per_unit_turnover,
            "frobenius":         result.frobenius,
            "spectral":          result.spectral,
            "stein":             result.stein,
            "precision":         result.precision,
        }
        cls._append(path, SHEET_SUMMARY, _SUMMARY_COLS, values)

    @classmethod
    def write_log(cls, path: Path, result: ScenarioResult) -> None:
        """1 ligne dans run_log."""
        values = {**result.key.to_dict(), "status": result.status, "duration_s": round(result.duration, 1) if result.duration else None, "error": result.error,}
        cls._append(path, SHEET_LOG, _LOG_COLS, values)



# Classe interne : extraction et calcul des métriques
class _Parser:
    """
    Classe contenant les méthodes d'extraction et de calcul des métriques depuis
    le dictionnaire retourné par full_evaluation(). Sépare la partie économique
    (backtests TE-min) et la partie statistique (pertes matricielles).

    Methods
    -------
    parse(eval_res, model_name, ann_factor) -> dict :
        Parse le dictionnaire complet de résultats pour un modèle donné.
    """

    @classmethod
    def parse(cls, eval_res: Dict[str, Any],model_name: str,ann_factor : int = 252) -> Dict[str, Any]:   
        out: Dict[str, Any] = {}

        # Fusionne les métriques économiques et les pertes matricielles dans un seul dict
        out.update(cls._eco(eval_res, model_name,ann_factor))
        out.update(cls._stats(eval_res, model_name))
        return out

    # Partie économique
    @classmethod
    def _eco(cls, eval_res: Dict[str, Any],model_name: str,ann_factor : int = 252) -> Dict[str, Any]:
        """
        Extrait et calcule les métriques économiques depuis les résultats de backtest.

        Récupère le BacktestResult du modèle, calcule les active returns,
        la TE ex-post rolling et les métriques agrégées via _agg_metrics().
        """
        # Navigue dans le dict de résultats pour trouver le BacktestResult du modèle
        bt = eval_res.get("economic", {}).get("bt_results", {}).get(model_name)

        # Si le modèle est absent des résultats, on retourne un dict vide
        if bt is None:
            return {}

        port_s  = _to_series(getattr(bt, "portfolio_returns",  None))
        bench_s = _to_series(getattr(bt, "benchmark_returns",  None))

        # Si l'une des deux séries est manquante, impossible de calculer les métriques
        if port_s is None or bench_s is None:
            return {}

        # Aligne les deux séries sur leur index commun pour éviter les décalages de dates
        idx      = port_s.index.intersection(bench_s.index)
        port_s   = port_s.reindex(idx)
        bench_s  = bench_s.reindex(idx)

        # Active returns journaliers : rendement portefeuille moins rendement benchmark
        active_s = port_s - bench_s

        # TE ex-post rolling : écart-type annualisé sur fenêtre glissante de ann_factor jours
        te_s = (active_s.rolling(ann_factor).std(ddof=1).mul(np.sqrt(ann_factor)))

        # Poids en séance récupérés pour le calcul du turnover dans _agg_metrics
        wif = getattr(bt, "weights_in_force", None)

        return {"port_returns": port_s, "bench_returns": bench_s,"active_returns": active_s,"te_expost":te_s, **cls._agg_metrics(port_s, bench_s, active_s, wif,ann_factor),}

    @staticmethod
    def _agg_metrics( port: pd.Series,bench: pd.Series,active: pd.Series, weights_in_force: Optional[pd.DataFrame] = None,ann_factor : int = 252) -> Dict[str, Optional[float]]:
        """
        Calcule toutes les métriques agrégées sur la période complète.

        Parameters
        ----------
        port / bench / active : pd.Series
            Rendements journaliers portefeuille, benchmark et actifs.
        weights_in_force : pd.DataFrame or None
            Poids en séance pour le calcul du turnover.
        ann_factor : int
            Facteur d'annualisation.
        """
        from scipy.stats import skew, kurtosis as kurt

        ann = ann_factor

        # Métriques de base sur les active returns
        te_d   = float(active.std(ddof=1))
        te_ann = te_d * np.sqrt(ann)
        mu_ann = float(active.mean()) * ann

        # IR : rendement actif annualisé divisé par la TE (évite la division par zéro)
        ir     = (mu_ann / te_ann) if te_ann > 1e-12 else np.nan

        # Performances cumulées sur toute la période
        port_cum = float((1 + port).prod()  - 1)
        bnch_cum = float((1 + bench).prod() - 1)

        # Max drawdown actif : pire chute de la valeur cumulée depuis son sommet
        cum_active = (1 + active).cumprod()
        roll_max   = cum_active.cummax()
        drawdown   = (cum_active - roll_max) / roll_max
        max_dd_active = _sf(float(drawdown.min()))

        # Skewness et kurtosis calculés uniquement si assez d'observations
        a = active.dropna().values
        skewness_active = _sf(float(skew(a)))        if len(a) > 3 else None
        kurtosis_active = _sf(float(kurt(a, fisher=True))) if len(a) > 3 else None

        # TE stability : écart-type des TE annuelles, mesure la régularité inter-annuelle
        active_ann = active.groupby(active.index.year).apply(lambda x: float(x.std(ddof=1) * np.sqrt(ann)) if len(x) > 5 else np.nan)
        te_stability = _sf(float(active_ann.std())) if len(active_ann.dropna()) > 1 else None

        # Pire et meilleure année de rendement actif cumulé
        ar_by_year = active.groupby(active.index.year).apply(lambda x: float((1 + x).prod() - 1))
        worst_year_active = _sf(float(ar_by_year.min())) if len(ar_by_year) > 0 else None
        best_year_active  = _sf(float(ar_by_year.max())) if len(ar_by_year) > 0 else None

        # TE bear/bull : TE calculée séparément selon le signe du benchmark
        bear_mask = bench < 0
        bull_mask = bench > 0
        te_ann_bear = _sf(float(active[bear_mask].std(ddof=1) * np.sqrt(ann))) \
                    if bear_mask.sum() > 5 else None
        te_ann_bull = _sf(float(active[bull_mask].std(ddof=1) * np.sqrt(ann))) \
                    if bull_mask.sum() > 5 else None

        # Turnover annualisé : somme des variations L1/2 sur les jours de rebalancement réels
        turnover_ann = None
        if weights_in_force is not None and isinstance(weights_in_force, pd.DataFrame):

            # Variation L1/2 par jour (one-way turnover)
            dw = weights_in_force.diff().abs().sum(axis=1) / 2.0

            # annualisation : moyenne des jours de rebal x 252, on prend les jours où il y a eu un vrai trade (dw > seuil)
            rebal_days = dw[dw > 1e-6]
            if len(rebal_days) > 0:
                n_days_total = len(weights_in_force.dropna(how="all"))
                n_years = n_days_total / ann
                turnover_ann = _sf(float(rebal_days.sum()) / n_years) if n_years > 0 else None

        # TE per unit turnover : efficacité de l'estimateur par unité de coût de rebalancement
        te_per_unit_turnover = None
        if turnover_ann is not None and turnover_ann > 1e-6 and te_ann is not None:
            te_per_unit_turnover = _sf(te_ann / turnover_ann)

        return {
            "te_ann":               _sf(te_ann),
            "te_daily":             _sf(te_d),
            "information_ratio":    _sf(ir),
            "active_mean_ann":      _sf(mu_ann),
            "port_cum":             _sf(port_cum),
            "bench_cum":            _sf(bnch_cum),
            "rel_cum_vs_bench":     _sf(port_cum - bnch_cum),
            "max_dd_active":        max_dd_active,
            "skewness_active":      skewness_active,
            "kurtosis_active":      kurtosis_active,
            "te_stability":         te_stability,
            "te_ann_bear":          te_ann_bear,
            "te_ann_bull":          te_ann_bull,
            "worst_year_active":    worst_year_active,
            "best_year_active":     best_year_active,
            "turnover_ann":         turnover_ann,
            "te_per_unit_turnover": te_per_unit_turnover,
        }

    # Partie stats 
    @staticmethod
    def _stats(eval_res: Dict[str, Any],
               model_name: str) -> Dict[str, Optional[float]]:
        """
        Extrait les pertes matricielles depuis la section 'simulation' des résultats.

        Retourne None pour chaque métrique si la section stats est absente,
        ce qui est le cas quand on lance uniquement le backtest économique.
        """
        try:
            ls  = eval_res["simulation"]["loss_summary"]
            row = ls.loc[model_name] if model_name in ls.index else pd.Series()
            return {
                "frobenius": _sf(row.get("frobenius")),
                "spectral": _sf(row.get("spectral")),
                "stein": _sf(row.get("stein")),
                "precision": _sf(row.get("precision")),
            }
        except Exception:
            return {"frobenius": None, "spectral": None,
                    "stein": None, "precision": None}



# Utilitaires
def _sf(x) -> Optional[float]:
    """safe float : convertit x en float valide, retourne None si NaN ou non-convertible."""

    try:
        v = float(x)

        # NaN n'est pas une valeur valide pour Excel, on retourne None à la place
        return None if np.isnan(v) else v
    
    except Exception:
        return None


def _to_series(x) -> Optional[pd.Series]:
    """Convertit x en pd.Series et supprime les NaN. Retourne None si vide."""

    if x is None:
        return None
    
    s = pd.Series(x).dropna()
    return s if len(s) > 0 else None


def _reindex_or_nan(s: Optional[pd.Series], idx: pd.Index) -> pd.Series:
    """Réindexe une série sur idx. Si s est None, retourne une série de NaN sur idx."""

    if s is None:
        # Aucune donnée disponible : on retourne une série de NaN pour ne pas planter les calculs
        return pd.Series(np.nan, index=idx)
    
    return s.reindex(idx)


# Interface publique : branche économique
class MonteCarloExporter:
    """
    Interface principale pour l'export des résultats du Monte Carlo économique.
    Gère le checkpoint (skip des scénarios déjà calculés), le buffer en mémoire
    et le flush par lots dans Excel.

    Deux modes d'écriture disponibles :
    - write_last() : écriture immédiate scénario par scénario (atomique).
    - write() + flush() : buffering en mémoire avec flush tous les N scénarios
      pour minimiser les I/O disque sur de grands Monte Carlo.

    Attributes
    ----------
    path : Path
        Chemin du fichier Excel de sortie.
    flush_every : int
        Nombre de scénarios bufferisés avant flush automatique.

    Methods
    -------
    already_done(key) -> bool :
        Retourne True si le scénario est déjà dans le checkpoint.
    build_result(key, eval_res, ann_factor, duration) -> ScenarioResult :
        Construit un ScenarioResult depuis le dictionnaire de full_evaluation().
    build_error(key, exc, duration) -> ScenarioResult :
        Construit un ScenarioResult d'erreur.
    write_last(result) -> None :
        Écriture immédiate dans les 3 feuilles Excel.
    write(result) -> None :
        Buffering + flush automatique.
    flush() -> None :
        Force l'écriture de tous les résultats bufferisés dans Excel.
    read_ts() / read_summary() / read_log() -> pd.DataFrame :
        Lecture des feuilles Excel pour analyse.
    """

    def __init__(self, path: str | Path,flush_every: int = 100) -> None:

        # Chemin d'accès au fichier Excel de sortie
        self.path = Path(path)
        self.flush_every = flush_every

        # Crée le fichier si absent, ou vérifie que les feuilles attendues sont présentes
        _Initializer.init(self.path)

        # Charge les scénarios déjà calculés pour pouvoir les skipper
        self._done: Set[ScenarioKey] = _Checkpoint.load(self.path)

        # Buffer en mémoire : les résultats s'accumulent ici avant d'être flushés
        self._buffer: List[ScenarioResult] = []
        logger.info(f"[Export] {len(self._done)} scénario(s) déjà calculé(s).")

    # Checkpoint
    def already_done(self, key: ScenarioKey) -> bool:
        """Retourne True si le scénario est déjà présent dans le checkpoint."""
        return key in self._done

    def n_done(self) -> int:
        """Retourne le nombre de scénarios déjà calculés."""
        return len(self._done)

    # Construction
    def build_result(self, key: ScenarioKey, eval_res: Dict[str, Any],ann_factor: int = 252, duration: Optional[float] = None) -> ScenarioResult:
                     
        parsed = _Parser.parse(eval_res, key.model_name)
        result = ScenarioResult(key=key, duration=duration)

        # Affecte chaque métrique parsée à l'attribut correspondant du ScenarioResult
        for k, v in parsed.items():
            if hasattr(result, k):
                setattr(result, k, v)
        return result

    def build_error(self, key: ScenarioKey, exc: Exception,duration: Optional[float] = None) -> ScenarioResult:
        """Construit un ScenarioResult d'erreur avec message tronqué à 250 caractères."""
        return ScenarioResult(  key=key, status="error",  error=str(exc)[:250], duration=duration,)

       
    # Écriture 
    def write_last(self, result: ScenarioResult) -> None:
        """Écrit ts_results (N lignes/dates), summary (1 ligne), log (1 ligne)."""
        
        if result.status == "success":
            _Writer.write_ts(self.path, result,result.ann_factor)
            
        _Writer.write_summary(self.path, result)
        _Writer.write_log(self.path, result)
        self._done.add(result.key)
        logger.info(f"[Export] Écrit : {result.key}")

    def write(self, result: ScenarioResult) -> None:
        """Ajoute au buffer. Flush automatique tous les flush_every scénarios."""

        self._buffer.append(result)
        self._done.add(result.key)
        logger.info(f"[Export] Bufferisé : {result.key} ({len(self._buffer)}/{self.flush_every})")

        # Déclenche le flush automatique quand le buffer est plein
        if len(self._buffer) >= self.flush_every:
            self.flush()

    def flush(self) -> None:
        """Écrit tous les résultats en attente dans Excel en une seule passe."""
        if not self._buffer:
            return

        # Ouvre le fichier une seule fois pour écrire tous les résultats bufferisés
        wb = load_workbook(self.path)

        for result in self._buffer:
            if result.status == "success":
                _Writer.write_ts_wb(wb, result, result.ann_factor)
            _Writer.write_summary_wb(wb, result)
            _Writer.write_log_wb(wb, result)

        # Sauvegarde une seule fois après avoir écrit tous les résultats
        wb.save(self.path)
        logger.info(f"[Export] Flush de {len(self._buffer)} scénarios → {self.path}")

        # Vide le buffer après le flush
        self._buffer.clear()

    # Lecture 
    def read_ts(self) -> pd.DataFrame:
        """Lit la feuille ts_results depuis le fichier Excel."""
        return pd.read_excel(self.path, sheet_name=SHEET_TS,      engine="openpyxl")

    def read_summary(self) -> pd.DataFrame:
        """Lit la feuille summary depuis le fichier Excel."""
        return pd.read_excel(self.path, sheet_name=SHEET_SUMMARY, engine="openpyxl")

    def read_log(self) -> pd.DataFrame:
        """Lit la feuille run_log depuis le fichier Excel."""
        return pd.read_excel(self.path, sheet_name=SHEET_LOG,     engine="openpyxl")
    



# Branche statistique : StatScenarioKey + StatMonteCarloExporter

# Noms des feuilles Excel pour la partie stats
STAT_SHEET_SUMMARY = "stat_summary"
STAT_SHEET_LOG     = "stat_log"

# Colonnes de la clé stats : identifient un scénario statistique de façon unique
_STAT_KEY_COLS = ["model_name", "dgp_type", "N_sim", "n_factors", "innovation", "p_ratio", "rho_B", "sigma_B", "seed",]

# Colonnes du summary stats : clé + une colonne par perte matricielle
_STAT_SUMMARY_COLS = _STAT_KEY_COLS + ["frobenius", "spectral", "stein", "precision"]

# Colonnes du log stats
_STAT_LOG_COLS = _STAT_KEY_COLS + ["status", "duration_s", "error"]


@dataclass(frozen=True)
class StatScenarioKey:
    """
    Classe contenant les paramètres qui identifient de façon unique un scénario
    de simulation statistique. Utilisée comme clé de checkpoint.

    Attributes
    ----------
    model_name : str
        Nom du modèle de covariance testé.
    dgp_type : str
        Type de DGP : 'static_oracle' ou 'factor_shock'.
    N_sim : int
        Taille de la matrice simulée (nombre d'actifs).
    n_factors : int
        Nombre de facteurs du DGP.
    innovation : str
        Type d'innovation : 'gaussian' ou 'student'.
    p_ratio : float or None
        Ratio p = N/T (static_oracle uniquement). None pour factor_shock.
    rho_B : float or None
        Persistance AR(1) des loadings (factor_shock uniquement). None sinon.
    sigma_B : float or None
        Volatilité des chocs sur les loadings. None sinon.
    seed : int
        Graine aléatoire du scénario.

    Methods
    -------
    to_dict() -> dict :
        Sérialise la clé en dictionnaire pour l'écriture dans Excel.
    """
    model_name : str
    dgp_type   : str             
    N_sim      : int
    n_factors  : int
    innovation : str             
    p_ratio    : Optional[float] 
    rho_B : Optional[float] 
    sigma_B : Optional[float] 
    seed : int             

    def to_dict(self) -> Dict[str, Any]:
        """Sérialise la clé en dictionnaire (pour l'écriture dans Excel)."""
        return asdict(self)

    def __str__(self) -> str:
        return (f"{self.model_name} | dgp={self.dgp_type} | " f"N={self.N_sim} | seed={self.seed}")


@dataclass
class StatScenarioResult:
    """
    Classe contenant les pertes matricielles et le statut d'un scénario statistique.

    Attributes
    ----------
    key : StatScenarioKey
        Identifiant du scénario.
    frobenius / spectral / stein / precision : float or None
        Pertes matricielles (Frobenius, spectrale, Stein, précision).
    status : str
        'success' ou 'error'.
    error : str or None
        Message d'erreur tronqué à 250 caractères.
    duration : float or None
        Durée d'exécution en secondes.
    """

    key : StatScenarioKey
    frobenius : Optional[float] = None
    spectral : Optional[float] = None
    stein : Optional[float] = None
    precision  : Optional[float] = None
    status : str             = "success"
    error : Optional[str]   = None
    duration   : Optional[float] = None


class _StatInitializer:
    """Classe contenant les méthodes de création et de vérification du fichier Excel stats."""

    SHEETS = {STAT_SHEET_SUMMARY: _STAT_SUMMARY_COLS, STAT_SHEET_LOG: _STAT_LOG_COLS,}

    @classmethod
    def init(cls, path: Path) -> None:
        if path.exists():
            cls._ensure_sheets(path)
        else:
            cls._create(path)

    @classmethod
    def _create(cls, path: Path) -> None:
        """Crée un nouveau fichier Excel avec les 2 feuilles statistiques."""
        wb = Workbook()

        # Supprime la feuille vide créée automatiquement par openpyxl
        wb.remove(wb.active)

        for name, cols in cls.SHEETS.items():
            ws = wb.create_sheet(name)
            ws.append(cols)
            _Styler.apply(ws, cols)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        logger.info(f"[StatExport] Fichier créé : {path}")

    @classmethod
    def _ensure_sheets(cls, path: Path) -> None:
        """Ajoute les feuilles manquantes dans un fichier existant."""
        wb = load_workbook(path)
        changed = False
        for name, cols in cls.SHEETS.items():
            if name not in wb.sheetnames:
                ws = wb.create_sheet(name)
                ws.append(cols)
                _Styler.apply(ws, cols)
                changed = True

        # Sauvegarde uniquement si au moins une feuille a été ajoutée
        if changed:
            wb.save(path)


class _StatCheckpoint:
    """Classe contenant les méthodes de lecture des scénarios statistiques déjà calculés."""

    @staticmethod
    def load(path: Path) -> Set[StatScenarioKey]:
        """Charge les clés existantes depuis stat_summary."""
        if not path.exists():
            return set()
        try:
            df = pd.read_excel(path, sheet_name=STAT_SHEET_SUMMARY, engine="openpyxl")
            return _StatCheckpoint._parse(df)
        except Exception as e:
            logger.warning(f"[StatCheckpoint] Lecture impossible : {e}")
            return set()

    @staticmethod
    def _parse(df: pd.DataFrame) -> Set[StatScenarioKey]:
        """Convertit les lignes du DataFrame en StatScenarioKey. Ignore les lignes malformées."""
        keys: Set[StatScenarioKey] = set()

        # Si les colonnes attendues ne sont pas toutes présentes, le fichier est incompatible
        if not set(_STAT_KEY_COLS).issubset(df.columns):
            return keys

        for _, r in df.iterrows():
            try:
                keys.add(StatScenarioKey(
                    model_name = str(r["model_name"]),
                    dgp_type   = str(r["dgp_type"]),
                    N_sim      = int(r["N_sim"]),
                    n_factors  = int(r["n_factors"]),
                    innovation = str(r["innovation"]),

                    # p_ratio, rho_B, sigma_B peuvent être NaN dans Excel si non applicable
                    p_ratio    = None if pd.isna(r["p_ratio"]) else float(r["p_ratio"]),
                    rho_B      = None if pd.isna(r["rho_B"])   else float(r["rho_B"]),
                    sigma_B    = None if pd.isna(r["sigma_B"]) else float(r["sigma_B"]),
                    seed       = int(r["seed"]),
                ))
            except Exception:

                # La ligne est malformée, on passe à la suivante sans planter
                pass
        return keys


class _StatWriter:
    """Classe contenant les méthodes d'écriture des résultats statistiques sur un workbook ouvert."""

    @staticmethod
    def write_summary_wb(wb, result: StatScenarioResult) -> None:
        """Écrit la ligne stat_summary sur un workbook ouvert."""
        ws = wb[STAT_SHEET_SUMMARY]
        d  = result.key.to_dict()
        row = [d.get(c) for c in _STAT_KEY_COLS] + [result.frobenius, result.spectral, result.stein, result.precision,]
        ws.append(row)

    @staticmethod
    def write_log_wb(wb, result: StatScenarioResult) -> None:
        """Écrit la ligne stat_log sur un workbook ouvert."""
        ws  = wb[STAT_SHEET_LOG]
        d   = result.key.to_dict()
        row = [d.get(c) for c in _STAT_KEY_COLS] + [result.status, round(result.duration, 2) if result.duration else None, result.error,]
        ws.append(row)



# Interface publique : branche statistique
class StatMonteCarloExporter:
    """
    Interface principale pour l'export des résultats du Monte Carlo statistique.
    Même logique que MonteCarloExporter : checkpoint, buffer en mémoire, flush.
    Deux feuilles uniquement (stat_summary + stat_log), pas de ts_results.

    Attributes
    ----------
    path : Path
        Chemin du fichier Excel de sortie.
    flush_every : int
        Nombre de scénarios bufferisés avant flush automatique.

    Methods
    -------
    already_done(key) -> bool :
        Retourne True si le scénario est déjà dans le checkpoint.
    build_result(key, loss_row, duration) -> StatScenarioResult :
        Construit un StatScenarioResult depuis la ligne de pertes.
    build_error(key, exc, duration) -> StatScenarioResult :
        Construit un StatScenarioResult d'erreur.
    write(result) -> None :
        Buffering + flush automatique.
    flush() -> None :
        Force l'écriture dans Excel.
    read_summary() -> pd.DataFrame :
        Lit la feuille stat_summary.
    """

    def __init__(self, path: str | Path, flush_every: int = 50) -> None:
        self.path        = Path(path)
        self.flush_every = flush_every

        # Crée le fichier si absent, ou vérifie que les feuilles attendues sont présentes
        _StatInitializer.init(self.path)

        # Charge les scénarios déjà calculés pour pouvoir les skipper
        self._done  : Set[StatScenarioResult] = _StatCheckpoint.load(self.path)

        # Buffer en mémoire : les résultats s'accumulent ici avant d'être flushés
        self._buffer: List[StatScenarioResult] = []
        logger.info(f"[StatExport] {len(self._done)} scénario(s) déjà calculé(s).")

    def already_done(self, key: StatScenarioKey) -> bool:
        """Retourne True si le scénario est déjà dans le checkpoint."""
        return key in self._done

    def n_done(self) -> int:
        """Retourne le nombre de scénarios déjà calculés."""
        return len(self._done)

    def build_result(self, key : StatScenarioKey, loss_row : Dict[str, Optional[float]], duration : Optional[float] = None,) -> StatScenarioResult:
        """
        Construit un StatScenarioResult depuis la ligne de pertes
        retournée par _run_one_scenario pour un modèle donné.
        """
        return StatScenarioResult(
            key       = key,

            # Convertit chaque perte en safe float pour éviter les NaN dans Excel
            frobenius = _sf(loss_row.get("frobenius")),
            spectral  = _sf(loss_row.get("spectral")),
            stein     = _sf(loss_row.get("stein")),
            precision = _sf(loss_row.get("precision")),
            duration  = duration,
        )

    def build_error(self, key : StatScenarioKey, exc : Exception, duration : Optional[float] = None,) -> StatScenarioResult:
        """Construit un StatScenarioResult d'erreur."""

        return StatScenarioResult(key = key, status = "error", error = str(exc)[:250], duration = duration,)

    def write(self, result: StatScenarioResult) -> None:
        """Ajoute au buffer. Flush automatique tous les flush_every scénarios."""
        self._buffer.append(result)
        self._done.add(result.key)

        # Déclenche le flush automatique quand le buffer est plein
        if len(self._buffer) >= self.flush_every:
            self.flush()

    def flush(self) -> None:
        """Écrit tous les résultats en attente dans Excel en une seule passe."""
        if not self._buffer:
            return

        # Ouvre le fichier une seule fois pour écrire tous les résultats bufferisés
        wb = load_workbook(self.path)
        for result in self._buffer:
            _StatWriter.write_summary_wb(wb, result)
            _StatWriter.write_log_wb(wb, result)
        wb.save(self.path)
        logger.info(f"[StatExport] Flush {len(self._buffer)} scénarios → {self.path}")

        # Vide le buffer après le flush
        self._buffer.clear()

    def read_summary(self) -> pd.DataFrame:
        """Lit la feuille stat_summary depuis le fichier Excel."""
        return pd.read_excel(self.path, sheet_name=STAT_SHEET_SUMMARY, engine="openpyxl")