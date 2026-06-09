"""
Export de l'inventaire de portefeuille au format Bloomberg PORT.
 
Ce fichier gère la génération du fichier Excel d'inventaire produit après chaque
backtest. Il exporte les poids du portefeuille à chaque date de rebalancement dans
un format compatible Bloomberg PORT, et produit deux feuilles supplémentaires :
une feuille de comparaison portefeuille vs benchmark (poids actifs), et une feuille
de statistiques par rebalancement (turnover, active share, rendements inter-rebal, etc.).


Classes
-------
PortInventoryConfig :
    Dataclass de configuration de l'export : nom du portefeuille, chemin de sortie, source des poids, mode de détection des dates de rebalancement, et options de mise en forme.
PortInventoryExporter :
    Classe principale qui extrait les snapshots de poids à chaque date de rebalancement depuis un BacktestResult, calcule les statistiques, et écrit le fichier Excel formaté.
 
Fonctions
---------
build_port_inventory :
    Fonction helper pour instancier et exécuter l'export directement après engine.run().
"""

from __future__ import annotations

import warnings
from dataclasses import dataclass, field
from pathlib import Path
from typing import Literal, Optional

import numpy as np
import pandas as pd


# Types
RebalanceDateMode = Literal["weights_in_force_change", "weights_close_change", "trades_dt",]
WeightSource = Literal["weights_in_force", "weights_close",]


# Config
@dataclass
class PortInventoryConfig:
    """
    Configuration de l'export inventaire Bloomberg PORT.
 
    Regroupe tous les paramètres nécessaires à l'export : identification du portefeuille,
    chemin de sortie, source des poids, options de filtrage et de mise en forme Excel.
 
    Attributes
    ----------
    ptf_name : str
        Nom du portefeuille tel qu'il apparaîtra dans la colonne 'ptf_name' du fichier Bloomberg PORT.
    output_path : Path
        Chemin complet du fichier .xlsx de sortie (créé automatiquement avec ses répertoires parents).
    enabled : bool
        Si False, l'export est désactivé et run_if_enabled() retourne None sans écrire de fichier.
    rebalance_date_mode : str
        Méthode de détection des dates de rebalancement
    weight_source : str
        Source des poids à exporter : 'weights_in_force' ou 'weights_close'.
    tol : float
        Seuil de tolérance numérique pour considérer un poids comme nul (défaut : 1e-6).
    drop_zeros : bool
        Si True, les actifs avec |poids| < tol sont exclus du snapshot exporté.
    include_cash : bool
        Si True, ajoute une ligne 'CASH' avec le poids de la poche cash dans chaque snapshot.
    cash_ticker : str
        Nom du ticker cash dans le fichier exporté (défaut : 'CASH').
    date_format : str
        Format de date utilisé dans le fichier Excel (défaut : '%Y-%m-%d').
    sheet_name : str
        Nom de la feuille principale dans le fichier Excel Bloomberg PORT.
    apply_lag : int
        Décalage en jours de trading appliqué pour récupérer les poids effectifs après la date de décision (alignement sur l'apply_lag du backtest).
    """
    
    ptf_name: str
    output_path: Path
    enabled: bool = True
    rebalance_date_mode: RebalanceDateMode = "weights_in_force_change"
    weight_source: WeightSource = "weights_in_force"
    tol: float = 1e-6
    drop_zeros: bool = True
    include_cash: bool = False
    cash_ticker: str = "CASH"
    date_format: str = "%Y-%m-%d"
    sheet_name: str = "port_inventory"
    apply_lag: int = 1 

    def __post_init__(self) -> None:
        """ Post-initialisation pour valider les paramètres et convertir output_path en Path si nécessaire."""
        self.output_path = Path(self.output_path)
    
    
class PortInventoryExporter:

    """
    Extrait l'inventaire du portefeuille à chaque date de rebalancement et l'exporte
    dans un fichier Excel au format Bloomberg PORT.
 
    Produit trois feuilles dans le fichier de sortie :
    - 'port_inventory' : poids par actif et par date de rebalancement (format Bloomberg PORT).
    - 'active_weights' : comparaison portefeuille vs benchmark (poids actifs) par date de rebal.
    - 'stats'          : statistiques agrégées par rebalancement (turnover, active share, HHI,
                         rendements inter-rebal, top surpondérations/sous-pondérations).
 
    Attributes
    ----------
    cfg : PortInventoryConfig
        Configuration de l'export.
 
    Methods
    -------
    run(result) -> pd.DataFrame :
        Extrait les snapshots, calcule les statistiques et écrit le fichier Excel. Retourne le DataFrame d'inventaire (feuille principale).
    run_if_enabled(result) -> pd.DataFrame or None :
        Appelle run() uniquement si cfg.enabled=True. Retourne None sinon.
    """

    
    # Noms de colonnes fixes
    COL_PTF  = "ptf_name"
    COL_DATE = "rebal_date"
    COL_TKR  = "ticker"
    COL_WGT  = "weight"
    
    def __init__(self, cfg: PortInventoryConfig) -> None:
        """ Initialise l'exporteur avec la configuration donnée."""
        self.cfg = cfg
        self._active_df: Optional[pd.DataFrame] = None  
        self._stats_df:  Optional[pd.DataFrame] = None 
    
    
    def run(self, result) -> pd.DataFrame:

        """
        Extrait l'inventaire et écrit le fichier xlsx.
 
        Itère sur toutes les dates de rebalancement du BacktestResult, construit
        un snapshot de poids pour chacune, calcule le comparatif vs benchmark
        et les statistiques globales, puis écrit le fichier Excel formaté.
 
        Parameters
        ----------
        result : BacktestResult
            Résultat de backtest produit par BacktestEngine.run().
 
        Returns
        -------
        pd.DataFrame
            DataFrame d'inventaire au format Bloomberg PORT (colonnes : ptf_name, rebal_date, ticker, weight).
        """

        # Recupere les dates de rebalancement
        rebal_dates = result.rebal_dates

        # Verification des dates de rebalancement
        if len(rebal_dates) == 0:
            warnings.warn("PortInventoryExporter: aucune date de rebalancement detectee. ",stacklevel=2,)
            return pd.DataFrame(columns=[self.COL_PTF, self.COL_DATE, self.COL_TKR, self.COL_WGT])


        #Initialisation des listes pour stocker les DataFrames de chaque snapshot et du comparatif actif
        frames        = []
        active_frames = []  

        # Iteration sur chaque date de rebal
        for dt in rebal_dates:

            # Construction du snapshot de poids pour la date de rebalancement dt
            snap = self._snapshot(result, dt)

            # Si le snapshot est valide (non None et non vide), on l'ajoute à la liste des frames à concaténer
            if snap is not None and not snap.empty:
                frames.append(snap)

            # Snapshot comparatif ptf vs bench
            snap_vs = self._snapshot_vs_bench(result, dt)
            if snap_vs is not None and not snap_vs.empty:
                active_frames.append(snap_vs)

        # Concatenation de tous les snapshots individuels en un seul DataFrame d'inventaire complet
        if not frames:
            inventory = pd.DataFrame(columns=[self.COL_PTF, self.COL_DATE, self.COL_TKR, self.COL_WGT])
        else:
            inventory = pd.concat(frames, ignore_index=True)

        # Stocke le df actif pour _write_xlsx
        self._active_df = pd.concat(active_frames, ignore_index=True) if active_frames else None

        # Calcul des statistiques par rebal
        self._stats_df = self._compute_stats(result) if self._active_df is not None else None

        # Ecriture du fichier Excel avec les trois feuilles : port_inventory, active_weights, stats
        self._write_xlsx(inventory, result=result)

        return inventory
    
    def run_if_enabled(self, result) -> Optional[pd.DataFrame]:
        """
        Appelle run() uniquement si cfg.enabled=True.
 
        Parameters
        ----------
        result : BacktestResult
            Résultat de backtest produit par BacktestEngine.run().
 
        Returns
        -------
        pd.DataFrame or None
            DataFrame d'inventaire si enabled=True, None sinon.
        """
        
        # Si l'export est désactivé, retourne None, sinon exécute l'export et retourne le DataFrame d'inventaire
        if not self.cfg.enabled:
            return None
        return self.run(result)
    

    
    def _snapshot(self, result, dt: pd.Timestamp) -> Optional[pd.DataFrame]:

        """
        Retourne un DataFrame long [ptf_name, rebal_date, ticker, weight]
        pour une date de rebalancement donnée.
 
        Récupère les poids depuis la source configurée (weights_in_force ou weights_close),
        applique le décalage apply_lag pour pointer sur les poids effectifs après décision,
        filtre les zéros si configuré, et formate le résultat.
 
        Parameters
        ----------
        result : BacktestResult
            Résultat de backtest.
        dt : pd.Timestamp
            Date de rebalancement pour laquelle extraire le snapshot.
 
        Returns
        -------
        pd.DataFrame or None
            DataFrame long avec les colonnes Bloomberg PORT, ou None si aucun poids disponible.
        """

        # Convertit dt en Timestamp pour être sûr du type
        dt = pd.Timestamp(dt)

        # Récupère la configuration pour faciliter l'accès aux paramètres
        cfg = self.cfg

        # Recuperation de la matrice de poids
        W = getattr(result, cfg.weight_source, None)

        # Fallback si pas de poids trouvés : essaye l'autre source de poids avant de lever une erreur
        if W is None or (hasattr(W, "empty") and W.empty):
            other = "weights_close" if cfg.weight_source == "weights_in_force" else "weights_in_force"
            W = getattr(result, other, None)
            if W is None or (hasattr(W, "empty") and W.empty):
                raise ValueError(f"result.{cfg.weight_source} et result.{other} sont tous les deux absents.")

        # Assure que W est un DataFrame et convertit les poids en float
        W = W.astype(float)

        #Si la date de rabalancement est dans l'index de W
        if dt in W.index:

            #On trouve la position de cette date dans l'index de W
            base_loc = W.index.get_loc(dt)

        #Si la date de rabalancement n'est pas dans l'index de W
        else:

            #On trouve la date la plus proche avant dt dans l'index de W
            prior = W.index[W.index <= dt]

            #Si aucune date avant dt n'est trouvée, on ne peut pas construire le snapshot, on retourne None
            if len(prior) == 0:
                return None
            
            #Sinon, on prend la position de la date la plus proche avant dt
            base_loc = W.index.get_loc(prior[-1])

        # Décalage de apply_lag jours ouvrés dans l'index réel de W
        shifted_loc = min(base_loc + int(cfg.apply_lag), len(W.index) - 1)
        dt_use = W.index[shifted_loc]

        # Recupere les poids au jour effectif
        row: pd.Series = W.loc[dt_use].fillna(0.0)

        # Si inclusion cash activée
        if cfg.include_cash:

            #Recupere la série cash et nav depuis result
            cash_ser = getattr(result, "cash", None)
            nav_ser  = getattr(result, "nav",  None)

            #Si les séries cash et nav sont présentes
            if cash_ser is not None and nav_ser is not None:

                # Si la date dt_use est présente dans les deux séries, on calcule le poids cash et on l'ajoute à la ligne
                if dt_use in cash_ser.index and dt_use in nav_ser.index:
                    c = float(cash_ser.loc[dt_use])
                    n = float(nav_ser.loc[dt_use])
                    row[cfg.cash_ticker] = (c / n) if n != 0 else 0.0

        # Filtrage des zeros
        if cfg.drop_zeros:
            row = row[row.abs() > float(cfg.tol)]

        # Si après filtrage il ne reste aucun actif, retourne None
        if row.empty:
            return None

        # Date affichée = date de décision (dt original, pas dt_use) pour rester cohérent avec la date de rebal Bloomberg PORT
        date_str = pd.Timestamp(dt_use).strftime(cfg.date_format)

        # Construction du DataFrame long avec les colonnes ptf_name, rebal_date, ticker, weight
        df = pd.DataFrame({self.COL_PTF: cfg.ptf_name, self.COL_DATE: date_str, self.COL_TKR: row.index.astype(str), self.COL_WGT: row.values.astype(float),})
        df = df.sort_values(self.COL_WGT, ascending=False).reset_index(drop=True)

        return df
    

    def _snapshot_vs_bench(self, result, dt: pd.Timestamp) -> Optional[pd.DataFrame]:
        """
        Retourne un DataFrame comparatif portefeuille vs benchmark pour une date de rebal.
 
        Construit un tableau avec pour chaque actif du benchmark : son poids dans le
        portefeuille, son poids dans le benchmark, et le poids actif (différence).
        Les actifs exclus du portefeuille (ptf_weight = 0) sont conservés pour permettre
        d'identifier les exclusions.
 
        Parameters
        ----------
        result : BacktestResult
            Résultat de backtest.
        dt : pd.Timestamp
            Date de rebalancement pour laquelle construire le comparatif.
 
        Returns
        -------
        pd.DataFrame or None
            DataFrame avec colonnes : rebal_date, ticker, ptf_weight, bench_weight, active_weight. Trié par active_weight décroissant.
        """

        # Convertit dt en Timestamp pour être sûr du type
        dt  = pd.Timestamp(dt)

        # Récupère la configuration pour faciliter l'accès aux paramètres
        cfg = self.cfg

        # Récupération des poids ptf depuis la source configurée
        W = getattr(result, cfg.weight_source, None)

        # Fallback si pas de poids trouvés : essaye l'autre source de poids avant de retourner None
        if W is None or (hasattr(W, "empty") and W.empty):
            other = "weights_close" if cfg.weight_source == "weights_in_force" else "weights_in_force"
            W = getattr(result, other, None)
        if W is None or (hasattr(W, "empty") and W.empty):
            return None

        # Assure que W est un DataFrame et convertit les poids en float
        W = W.astype(float)

        # Trouve la date

        #Si la date de rabalancement est dans l'index de W
        if dt in W.index:

            # On trouve la position de cette date dans l'index de W
            base_loc = W.index.get_loc(dt)

        # Si la date de rabalancement n'est pas dans l'index de W
        else:

            # On trouve la date la plus proche avant dt dans l'index de W
            prior = W.index[W.index <= dt]

            # Si aucune date avant dt n'est trouvée, on ne peut pas construire le snapshot, on retourne None
            if len(prior) == 0:
                return None
            
            # Sinon, on prend la position de la date la plus proche avant dt
            base_loc = W.index.get_loc(prior[-1])

        # Décalage de apply_lag jours ouvrés dans l'index réel de W
        shifted_loc = min(base_loc + int(cfg.apply_lag), len(W.index) - 1)
        dt_use = W.index[shifted_loc]

        # Recupere les poids ptf au jour effectif
        ptf_row: pd.Series = W.loc[dt_use].fillna(0.0)

        # Recupere les poids bench au jour effectif
        B = getattr(result, "bench_weights_in_force", None)

        # Fallback si pas de poids bench trouvés : essaye bench_weights_close avant de retourner None
        if B is None or (hasattr(B, "empty") and B.empty):
            B = getattr(result, "bench_weights_close", None)
        if B is None or (hasattr(B, "empty") and B.empty):
            return None

        # Assure que B est un DataFrame et convertit les poids en float
        B = B.astype(float)

        # Si la date de rabalancement n'est pas dans l'index de B, on prend la date la plus proche avant dt_use
        bench_row: pd.Series = B.loc[dt_use] if dt_use in B.index else B.iloc[shifted_loc]

        # Remplit les NaN par 0 pour éviter les problèmes de calcul d'écart actif
        bench_row = bench_row.fillna(0.0)

        # Liste de tous les tickers présents dans le bench avec poids significatif (> tol)
        all_tickers = bench_row[bench_row.abs() > float(cfg.tol)].index.tolist()

        # Ajoute les actifs ptf non présents dans bench (cas rare)
        ptf_only = [t for t in ptf_row[ptf_row.abs() > float(cfg.tol)].index if t not in all_tickers]
        all_tickers = all_tickers + ptf_only

        # Construire le DataFrame comparatif ptf vs bench
        ptf_w   = ptf_row.reindex(all_tickers).fillna(0.0)
        bench_w = bench_row.reindex(all_tickers).fillna(0.0)
        active_w = ptf_w - bench_w

        # Construction du DataFrame avec les colonnes : rebal_date, ticker, ptf_weight, bench_weight, active_weight
        df = pd.DataFrame({"ticker": all_tickers,"ptf_weight": ptf_w.values, "bench_weight":  bench_w.values, "active_weight": active_w.values,})

        # Tri : surpondérations en haut, sous-pondérations en bas
        df = df.sort_values("active_weight", ascending=False).reset_index(drop=True)

        # Ajoute la date de rebal comme colonne info
        df.insert(0, "rebal_date", dt_use.strftime(cfg.date_format))

        return df


    def _compute_stats(self, result) -> Optional[pd.DataFrame]:
        """
        Calcule les statistiques d'allocation par date de rebalancement.
 
        Pour chaque date de rebalancement, calcule : nombre de positions actives,
        nombre d'actifs exclus,  turnover L1/2 vs rebal précédent, active share,
        HHI de concentration portefeuille et benchmark, top surpondérations/sous-pondérations,
        et rendements cumulés sur la période inter-rebalancement.
 
        Parameters
        ----------
        result : BacktestResult
            Résultat de backtest (utilisé pour récupérer les rendements inter-rebal).
 
        Returns
        -------
        pd.DataFrame or None
            DataFrame avec une ligne par date de rebalancement et les colonnes :
        """
         
        # Si le DataFrame actif n'est pas disponible, on ne peut pas calculer les stats, on retourne None
        if self._active_df is None or self._active_df.empty:
            return None

        # Récupère la configuration pour faciliter l'accès aux paramètres
        cfg = self.cfg

        # Initialisation de la liste pour stocker les résultats de chaque date de rebalancement
        rows = []

        # Récupération des séries de rendements depuis result
        port_rets = getattr(result, "portfolio_returns", None)
        bench_rets = getattr(result, "benchmark_returns", None)

        #Récupération des poids depuis la source configurée pour calculer le turnover
        W = getattr(result, cfg.weight_source, None)
        if W is None:
            W = getattr(result, "weights_in_force", None)

        # Tri des dates de rebalancement pour assurer l'ordre chronologique
        rebal_dates_sorted = sorted(self._active_df["rebal_date"].unique())

        #Iteration sur chaque date de rebalancement pour calculer les statistiques
        for i, dt_str in enumerate(rebal_dates_sorted):

            # Sous-DataFrame pour la date de rebalancement courante
            sub = self._active_df[self._active_df["rebal_date"] == dt_str].copy()

            # DataFrame des actifs investis (|poids| > tol) pour calculer les métriques de sur/sous-pondération
            invested = sub[sub["ptf_weight"].abs() > float(cfg.tol)].copy()

            # Calcul des statistiques globales pour la date de rebalancement courante
            n_positions = int((sub["ptf_weight"].abs() > float(cfg.tol)).sum())
            n_excluded = int((sub["ptf_weight"].abs() <= float(cfg.tol)).sum())
            sum_ptf = float(sub["ptf_weight"].sum())
            sum_bench = float(sub["bench_weight"].sum())

            # Calcul de l'active share : somme des poids actifs absolus divisée par 2
            active_share = float(sub["active_weight"].abs().sum() / 2.0)

            # Calcul du HHI (Herfindahl-Hirschman Index) pour mesurer la concentration du portefeuille et du benchmark
            hhi_ptf   = float((sub["ptf_weight"] ** 2).sum())
            hhi_bench = float((sub["bench_weight"] ** 2).sum())

            #Initialisation du turnover à NaN
            turnover = np.nan

            #Si la matrice de poids W est disponible et que ce n'est pas la première date de rebalancement
            if W is not None and i > 0:
                # Sous-DataFrame de la date de rebalancement précédente
                dt_prev_str = rebal_dates_sorted[i - 1]
                sub_prev = self._active_df[self._active_df["rebal_date"] == dt_prev_str]

                # Si le snapshot précédent n'est pas vide
                if not sub_prev.empty:

                    # Recupère les tickers et poids des deux snapshots,
                    tickers_now = set(sub["ticker"])
                    tickers_prev = set(sub_prev["ticker"])

                    # Union de tous les tickers présents dans les deux snapshots pour calculer le turnover
                    all_t = tickers_now | tickers_prev

                    # Recupère les poids ptf pour tous les tickers dans les deux snapshots, en remplissant les absents par 0
                    w_now  = sub.set_index("ticker")["ptf_weight"].reindex(all_t).fillna(0.0)
                    w_prev = sub_prev.set_index("ticker")["ptf_weight"].reindex(all_t).fillna(0.0)

                    # Calcul du turnover L1/2 : somme des différences absolues divisée par 2
                    turnover = float((w_now - w_prev).abs().sum() / 2.0)

            # Si le DataFrame des actifs investis est vide, on initialise les métriques de sur/sous-pondération à des valeurs par défaut 
            if invested.empty:
                max_over_ticker  = ""
                max_over_weight  = np.nan
                max_under_ticker = ""
                max_under_weight = np.nan
                top3_over        = ""
                top3_under       = ""

            #Sinon, on calcule les métriques de sur/sous-pondération à partir du DataFrame des actifs investis
            else:
                inv_sorted = invested.sort_values("active_weight", ascending=False)

                # Plus grande surpondération
                max_over_row = inv_sorted.iloc[0]
                max_over_ticker = str(max_over_row["ticker"])
                max_over_weight = float(max_over_row["active_weight"])

                # Plus grande sous-pondération
                min_over_row = inv_sorted.iloc[-1]
                max_under_ticker = str(min_over_row["ticker"])
                max_under_weight = float(min_over_row["active_weight"])

                # Top 3 surpondérations
                top3_over = " | ".join(f"{r['ticker']}:{r['active_weight']:+.2%}" for _, r in inv_sorted.head(3).iterrows())

                # Top 3 sous-pondérations
                top3_under = " | ".join(f"{r['ticker']}:{r['active_weight']:+.2%}" for _, r in inv_sorted.tail(3).iloc[::-1].iterrows())
                

            # Initialisation des rendements à NaN
            port_ret_val  = np.nan
            bench_ret_val = np.nan
            active_ret_val = np.nan

            #Si les séries de rendements du portefeuille et du benchmark sont disponibles
            if port_rets is not None and bench_rets is not None:

                try:
                    # Dates de debut
                    dt_curr = pd.Timestamp(dt_str)

                    # Date de fin = date du prochain rebalancement ou dernière date des rendements si c'est le dernier rebalancement
                    if i + 1 < len(rebal_dates_sorted):
                        dt_next = pd.Timestamp(rebal_dates_sorted[i + 1])
                    else:
                        dt_next = port_rets.index[-1]

                    # Rendement cumulé sur la période inter-rebal
                    mask = (port_rets.index > dt_curr) & (port_rets.index <= dt_next)

                    # Si il y a des rendements dans la période inter-rebal, on calcule les rendements cumulés du portefeuille et du benchmark, ainsi que l'écart actif
                    if mask.any():
                        port_ret_val = float((1 + port_rets[mask]).prod() - 1)
                        bench_ret_val = float((1 + bench_rets[mask]).prod() - 1)
                        active_ret_val = port_ret_val - bench_ret_val

                #Passe s'il y a une erreur dans le calcul des rendements
                except Exception:
                    pass
            
            # Ajoute une ligne au résultat avec toutes les statistiques calculées pour la date de rebalancement courante
            rows.append({
                "rebal_date": dt_str,
                "n_positions": n_positions,
                "n_excluded": n_excluded,
                "turnover": turnover,
                "active_share": active_share,
                "sum_ptf": sum_ptf,
                "sum_bench": sum_bench,
                "hhi_ptf": hhi_ptf,
                "hhi_bench": hhi_bench,
                "max_over_ticker": max_over_ticker,
                "max_over_weight": max_over_weight,
                "max_under_ticker": max_under_ticker,
                "max_under_weight": max_under_weight,
                "top3_over": top3_over,
                "top3_under": top3_under,
                "port_ret": port_ret_val,
                "bench_ret": bench_ret_val,
                "active_ret": active_ret_val,
            })

        return pd.DataFrame(rows)


    def _style_stats_sheet(self, writer) -> None:
        """
        Applique le style à la feuille 'stats'.
 
        Header bleu marine, colonnes de rendements et de turnover en format %,
        coloration conditionnelle du rendement actif (vert = positif, rouge = négatif),
        fond orange pour les rebalancements avec un turnover élevé (> 20%).
 
        Parameters
        ----------
        writer : pd.ExcelWriter
            Writer openpyxl en cours d'écriture.
        """

        #Import des styles openpyxl pour la mise en forme conditionnelle
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # initialisation de la feuille stats
        ws = writer.book["stats"]

        # Styles
        header_fill = PatternFill("solid", fgColor="1F497D")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border_side = Side(style="thin", color="AAAAAA")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side,)
        alt_fill     = PatternFill("solid", fgColor="EBF1F8")
        orange_fill  = PatternFill("solid", fgColor="FCE4D6")   # turnover élevé

        # Style header
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Mapping des lettres de colonnes pour identifier les colonnes par nom
        col_letters = {c.value: c.column_letter for c in ws[1]}

        # Colonnes en format %
        pct_cols = {"turnover", "active_share", "max_over_weight", "max_under_weight",  "port_ret", "bench_ret", "active_ret", "hhi_ptf", "hhi_bench",}

        # Colonnes en format nombre entier
        int_cols = {"n_positions", "n_excluded"}

        #Itération sur les lignes pour appliquer les styles conditionnels
        for r_idx, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):

            # Récupère turnover et active_ret pour coloration
            turnover_val   = 0.0

            #Itération sur les cellules de la ligne
            for cell in row_cells:

                # Identifie le nom de la colonne à partir de la lettre de colonne
                col_name = {v: k for k, v in col_letters.items()}.get(cell.column_letter, "")

                #Si la colonne est turnover ou active_ret, on essaie de convertir la valeur en float pour l'utiliser dans la coloration conditionnelle
                if col_name == "turnover" and cell.value is not None:
                    try: turnover_val = float(cell.value)
                    except: pass

                #Si la colonne est active_ret, on essaie de convertir la valeur en float pour l'utiliser dans la coloration conditionnelle
                if col_name == "active_ret" and cell.value is not None:
                    try: active_ret_val = float(cell.value)
                    except: pass

            #Iteration sur les cellules de la ligne pour appliquer les styles conditionnels en fonction des valeurs de turnover et active_ret
            for cell in row_cells:

                # Identifie le nom de la colonne à partir de la lettre de colonne
                cell.border = thin_border
                col_name = {v: k for k, v in col_letters.items()}.get(cell.column_letter, "")

                # Format numérique
                if col_name in pct_cols:
                    cell.number_format = "0.00%"
                elif col_name in int_cols:
                    cell.number_format = "0"
                elif col_name in {"sum_ptf", "sum_bench"}:
                    cell.number_format = "0.0000"

                # Coloration active_ret
                if col_name == "active_ret":
                    try:
                        v = float(cell.value) if cell.value is not None else 0.0
                    except: v = 0.0
                    if v > 1e-6:
                        cell.fill = PatternFill("solid", fgColor="C6EFCE")
                        cell.font = Font(color="276221")
                    elif v < -1e-6:
                        cell.fill = PatternFill("solid", fgColor="FFC7CE")
                        cell.font = Font(color="9C0006")
                    else:
                        cell.fill = alt_fill if r_idx % 2 == 0 else PatternFill()

                # Coloration turnover élevé (> 20%)
                elif col_name == "turnover" and turnover_val > 0.20:
                    cell.fill = orange_fill

                else:
                    cell.fill = alt_fill if r_idx % 2 == 0 else PatternFill()

        # Autosize
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10,)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        ws.freeze_panes = "A2"



    def _write_xlsx(self, df: pd.DataFrame, result=None) -> None:
        """
        Écrit le fichier Excel avec les trois feuilles formatées.
 
        Crée le répertoire parent si absent, puis écrit les feuilles
        'port_inventory', 'active_weights' et 'stats' avec leurs styles respectifs.
 
        Parameters
        ----------
        df : pd.DataFrame
            DataFrame principal de l'inventaire Bloomberg PORT.
        result : BacktestResult or None
            Résultat de backtest.
        """
        
        # Assure que le répertoire de sortie existe
        out = self.cfg.output_path
        out.parent.mkdir(parents=True, exist_ok=True)

        # Ecriture du fichier Excel avec openpyxl pour permettre la mise en forme avancée
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=self.cfg.sheet_name)

            # Feuille active_weights : poids ptf + bench + écart actif
            if result is not None and self._active_df is not None and not self._active_df.empty:
                self._active_df.to_excel(writer, index=False, sheet_name="active_weights")
                self._style_active_sheet(writer)

            # Feuille stats : métriques par date de rebal
            if self._stats_df is not None and not self._stats_df.empty:
                self._stats_df.to_excel(writer, index=False, sheet_name="stats")
                self._style_stats_sheet(writer)

            # Style de la feuille principale port_inventory
            ws = writer.book[self.cfg.sheet_name]
    
            # Style header
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            header_fill  = PatternFill("solid", fgColor="1F497D")
            header_font  = Font(bold=True, color="FFFFFF", size=11)
            border_side  = Side(style="thin", color="AAAAAA")
            thin_border  = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

            # Style du header
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
    
            # Format weight column (4 décimales)
            col_letters = {c.value: c.column_letter for c in ws[1]}
            wgt_letter = col_letters.get(self.COL_WGT)

            # Si la colonne de poids est trouvée, applique le format numérique à 4 décimales
            if wgt_letter:
                for row_cells in ws.iter_rows(min_row=2):
                    for cell in row_cells:
                        if cell.column_letter == wgt_letter:
                            cell.number_format = "0.0000"
    
            # Alternating rows
            alt_fill = PatternFill("solid", fgColor="EBF1F8")
            for r_idx, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
                fill = alt_fill if r_idx % 2 == 0 else PatternFill()
                for cell in row_cells:
                    cell.fill   = fill
                    cell.border = thin_border
    
            # Autosize columns
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value is not None),  default=10,)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    
            # Freeze header row
            ws.freeze_panes = "A2"


    def _style_active_sheet(self, writer) -> None:
        """
        Applique le style à la feuille 'active_weights'.
 
        Header bleu marine, colonnes poids en format %, coloration conditionnelle
        des poids actifs (vert = surpondération, rouge = sous-pondération),
        fond gris pour les actifs exclus du portefeuille (ptf_weight = 0).
 
        Parameters
        ----------
        writer : pd.ExcelWriter
            Writer openpyxl en cours d'écriture.
        """

        # Import des styles openpyxl pour la mise en forme conditionnelle
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.styles.differential import DifferentialStyle
        from openpyxl.formatting.rule import CellIsRule

        # Initialisation de la feuille active_weights
        ws = writer.book["active_weights"]

        # Styles
        header_fill = PatternFill("solid", fgColor="1F497D")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border_side = Side(style="thin", color="AAAAAA")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side,)

        # Style header
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border

        # Repère les colonnes par leur nom
        col_letters = {c.value: c.column_letter for c in ws[1]}

        # Format % pour les colonnes de poids
        pct_format = "0.00%"
        for col_name in ["ptf_weight", "bench_weight", "active_weight"]:
            letter = col_letters.get(col_name)
            if letter:
                for row_cells in ws.iter_rows(min_row=2):
                    for cell in row_cells:
                        if cell.column_letter == letter:
                            cell.number_format = pct_format

        # Alternating rows + mise en évidence actifs exclus (ptf_weight = 0)
        alt_fill      = PatternFill("solid", fgColor="EBF1F8")
        excluded_fill = PatternFill("solid", fgColor="E8E8E8")  # gris = exclu du ptf
        ptf_letter    = col_letters.get("ptf_weight")

        # Itération sur les lignes pour appliquer les styles conditionnels
        for r_idx, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):

            # Détecte si ptf_weight = 0 (actif exclu)
            ptf_val = 0.0
            for cell in row_cells:
                if cell.column_letter == ptf_letter and cell.value is not None:
                    try:
                        ptf_val = float(cell.value)
                    except (TypeError, ValueError):
                        ptf_val = 0.0

            # Si ptf_weight est nul (exclu), fond gris, sinon alternance de blanc et bleu clair
            is_excluded = abs(ptf_val) < 1e-8
            base_fill   = excluded_fill if is_excluded else (alt_fill if r_idx % 2 == 0 else PatternFill())

            # Itération sur les cellules de la ligne pour appliquer les styles conditionnels
            for cell in row_cells:
                cell.border = thin_border

                # Si la cellule n'est pas dans la colonne active_weight, on applique le style de base (alternance + exclusion)
                if cell.column_letter != col_letters.get("active_weight"):
                    cell.fill = base_fill
                
                # Sinon on applique la coloration conditionnelle sur le poids actif
                else:

                    # Coloration conditionnelle poids actif
                    try:
                        v = float(cell.value) if cell.value is not None else 0.0
                    except (TypeError, ValueError):
                        v = 0.0
                    if v > 1e-6:
                        cell.fill = PatternFill("solid", fgColor="C6EFCE")   # vert clair
                        cell.font = Font(color="276221")
                    elif v < -1e-6:
                        cell.fill = PatternFill("solid", fgColor="FFC7CE")   # rouge clair
                        cell.font = Font(color="9C0006")
                    else:
                        cell.fill = base_fill

        # Autosize colonnes
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None),  default=10,)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        ws.freeze_panes = "A2"




#HELPER
def build_port_inventory(result,*, ptf_name: str, output_path: Path | str,  enabled: bool = True, rebalance_date_mode: RebalanceDateMode = "weights_in_force_change",
                         weight_source: WeightSource = "weights_in_force", drop_zeros: bool = True, tol: float = 1e-6, include_cash: bool = False,) -> Optional[pd.DataFrame]:
        
    """
    Fonction helper pour instancier et exécuter l'export directement après engine.run().
 
    Crée une PortInventoryConfig avec les paramètres fournis, instancie un
    PortInventoryExporter et appelle run_if_enabled() en une seule ligne.
 
    Parameters
    ----------
    result : BacktestResult
        Résultat de backtest produit par BacktestEngine.run().
    ptf_name : str
        Nom du portefeuille dans le fichier Bloomberg PORT.
    output_path : Path or str
        Chemin de sortie du fichier Excel.
    enabled : bool
        Si False, l'export est désactivé et la fonction retourne None.
    rebalance_date_mode : str
        Mode de détection des dates de rebalancement.
    weight_source : str
        Source des poids à exporter.
    drop_zeros : bool
        Si True, les actifs avec poids nul sont exclus.
    tol : float
        Seuil de tolérance numérique pour les zéros.
    include_cash : bool
        Si True, ajoute une ligne CASH dans chaque snapshot.
 
    Returns
    -------
    pd.DataFrame or None
        DataFrame d'inventaire si enabled=True, None sinon.
    """
    # Crée la configuration et exécute l'export
    cfg = PortInventoryConfig(ptf_name=ptf_name, output_path=Path(output_path), enabled=enabled, rebalance_date_mode=rebalance_date_mode,
                              weight_source=weight_source, drop_zeros=drop_zeros, tol=tol, include_cash=include_cash,)
    
    # Instancie le PortInventoryExporter avec la configuration
    exporter = PortInventoryExporter(cfg)
    
    return exporter.run_if_enabled(result)